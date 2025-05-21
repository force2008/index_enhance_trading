import pandas as pd
import backtrader as bt
import logging
import numpy as np
import talib
import os
from joblib import Parallel, delayed
import openpyxl
from openpyxl.styles import numbers
import plotly.graph_objects as go
import statsmodels.api as sm

# 设置日志
logger = logging.getLogger()
logger.setLevel(logging.INFO)
log_file = 'data_clean/backtest1.log'
os.makedirs(os.path.dirname(log_file), exist_ok=True)
if os.path.exists(log_file):
    os.remove(log_file)
file_handler = logging.FileHandler(log_file, encoding='utf-8')
file_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# 加载符号和指数权重
cache_file = "data_clean/csi500_data_cache.pkl"
if not os.path.exists(cache_file):
    raise FileError(f"Cache file {cache_file} not found. Run data_download_clean.py first.")
cached_data = pd.read_pickle(cache_file)
symbols = cached_data['symbols']
index_weights = cached_data['index_weights']

# 因子层：计算多因子
column_mapping = {
    "净资产收益率(%)": "roe",
    "主营业务收入增长率(%)": "revenue_growth",
    "资产负债率(%)": "debt_ratio",
    "净利润同比增长(%)": "net_profit_growth",
    "总资产增长率(%)": "asset_growth",
    "销售毛利率(%)": "gross_margin",
    "流动资产周转率(次)": "wc_turnover",
    "资产的经营现金流量回报率(%)": "cash_flow_to_returns",
}

def calculate_factors(symbol):
    try:
        # 加载数据
        stock_data = pd.read_csv(f"data_clean/daily/stock_{symbol}_cleaned.csv", parse_dates=['date'], index_col='date')
        pe_pb_data = pd.read_csv(f"data_clean/pe_pb/stock_{symbol}_pe_pb_cleaned.csv", parse_dates=['date'], index_col='date')
        financial_data = pd.read_csv(f"data_clean/financial/stock_{symbol}_financial_cleaned.csv", parse_dates=['date'], index_col='date')
        dividend_data = pd.read_csv(f"data_clean/dividend/stock_{symbol}_dividend_cleaned.csv", parse_dates=['date'], index_col='date')
        index_data = pd.read_csv("data_clean/index/csi500_index.csv", parse_dates=['date'], index_col='date')

        # 数据完整性检查
        if (stock_data['close'].isna().all() or (stock_data['close'] <= 0).all() or stock_data['close'].lt(0).any() or
            'pe' not in pe_pb_data.columns or pe_pb_data['pe'].isna().all() or
            'pb' not in pe_pb_data.columns or pe_pb_data['pb'].isna().all() or
            '净资产收益率(%)' not in financial_data.columns or financial_data['净资产收益率(%)'].isna().all() or
            'dividend_yield' not in dividend_data.columns or dividend_data['dividend_yield'].isna().all()):
            logging.warning(f"股票 {symbol} 数据不完整")
            return pd.DataFrame()

        factors = pd.DataFrame(index=stock_data.index)
        financial_data = financial_data.rename(columns=column_mapping)

        # 1. 价值因子
        factors['pe'] = pe_pb_data['pe'].ffill().bfill().fillna(pe_pb_data['pe'].mean()).fillna(20)
        factors['pb'] = pe_pb_data['pb'].ffill().bfill().fillna(pe_pb_data['pb'].mean()).fillna(2)
        factors['ps'] = financial_data.get('总收入', stock_data['close'] * 100).div(stock_data['close']).ffill().bfill().fillna(20)
        factors['dividend_yield'] = dividend_data['dividend_yield'].ffill().bfill().fillna(0)

        # 2. 成长因子
        factors['roe'] = financial_data['roe'].ffill().bfill().fillna(financial_data['roe'].mean()).fillna(0)
        factors['revenue_growth'] = financial_data['revenue_growth'].ffill().bfill().fillna(financial_data['revenue_growth'].mean()).fillna(0)
        factors['net_profit_growth'] = financial_data.get('net_profit_growth', 0).ffill().bfill().fillna(0)
        factors['asset_growth'] = financial_data.get('asset_growth', 0).ffill().bfill().fillna(0)

        # 3. 动量因子
        factors['short_momentum'] = stock_data['close'].pct_change(20).fillna(0)
        factors['long_momentum'] = stock_data['close'].pct_change(120).fillna(0)
        factors['reversal'] = -stock_data['close'].pct_change(5).fillna(0)

        # 4. 技术因子
        factors['rsi'] = pd.Series(talib.RSI(stock_data['close'], timeperiod=14), index=stock_data.index).fillna(50)
        upper, middle, lower = talib.BBANDS(stock_data['close'], timeperiod=20)
        factors['bb_width'] = ((upper - lower) / middle).fillna(0)
        macd, signal, _ = talib.MACD(stock_data['close'], fastperiod=12, slowperiod=26, signalperiod=9)
        factors['macd_diff'] = (pd.Series(macd, index=stock_data.index) - pd.Series(signal, index=stock_data.index)).fillna(0)

        # 5. 波动率因子
        factors['raw_volatility'] = stock_data['close'].pct_change().rolling(20, min_periods=10).std().fillna(0.1)
        if (factors['raw_volatility'] < 0).any():
            logging.error(f"股票 {symbol} 包含负原始波动率")
            return pd.DataFrame()
        factors['low_vol'] = -factors['raw_volatility']
        market_returns = index_data['close'].pct_change().fillna(0)
        stock_returns = stock_data['close'].pct_change().fillna(0)
        idio_vol = []
        for t in stock_data.index:
            window = pd.DataFrame({
                'stock': stock_returns.loc[:t].tail(60),
                'market': market_returns.loc[:t].tail(60)
            }).dropna()
            if len(window) < 20:
                idio_vol.append(0)
                continue
            X = sm.add_constant(window['market'])
            model = sm.OLS(window['stock'], X).fit()
            residuals = model.resid
            idio_vol.append(residuals.std())
        factors['idio_vol'] = pd.Series(idio_vol, index=stock_data.index).fillna(0.1)

        # 6. 质量因子
        factors['debt_ratio'] = financial_data['debt_ratio'].ffill().bfill().fillna(financial_data['debt_ratio'].mean()).fillna(50)
        factors['gross_margin'] = financial_data.get('gross_margin', 0).ffill().bfill().fillna(0)
        factors['wc_turnover'] = financial_data['wc_turnover'].ffill().bfill().fillna(financial_data['wc_turnover'].mean()).fillna(0)
        
        # Handle cash_flow_to_assets
        if 'cash_flow_to_assets' in financial_data.columns:
            cash_flow_data = financial_data['cash_flow_to_assets']
            if isinstance(cash_flow_data, pd.DataFrame):
                logging.warning(f"股票 {symbol} cash_flow_to_assets 包含多个列，选取第一个有效列")
                cash_flow_data = cash_flow_data.iloc[:, 0]  # Select first column if DataFrame
            factors['cash_flow_to_assets'] = cash_flow_data.ffill().bfill().fillna(cash_flow_data.mean()).fillna(0)
        else:
            logging.warning(f"股票 {symbol} 缺少 cash_flow_to_assets 列，使用默认值 0")
            factors['cash_flow_to_assets'] = 0

        # 7. 市场因子
        factors['turnover_rate'] = (stock_data['volume'] / stock_data['volume'].mean()).rolling(20).mean().fillna(0)

        # 因子列表
        factor_columns = [
            'pe', 'pb', 'ps', 'dividend_yield',
            'roe', 'revenue_growth', 'net_profit_growth', 'asset_growth',
            'short_momentum', 'long_momentum', 'reversal',
            'rsi', 'bb_width', 'macd_diff',
            'low_vol', 'idio_vol',
            'debt_ratio', 'gross_margin', 'wc_turnover', 'cash_flow_to_assets',
            'turnover_rate'
        ]

        # 标准化因子
        for col in factor_columns:
            mean = factors[col].mean()
            std = factors[col].std()
            factors[col] = factors[col] - mean if pd.isna(std) or std == 0 else (factors[col] - mean) / std

        # 等权计算 score
        factors['score'] = (
            -factors['pe'] - factors['pb'] - factors['ps'] + factors['dividend_yield'] +
            factors['roe'] + factors['revenue_growth'] + factors['net_profit_growth'] + factors['asset_growth'] +
            factors['short_momentum'] + factors['long_momentum'] + factors['reversal'] -
            factors['rsi'] - factors['bb_width'] + factors['macd_diff'] +
            factors['low_vol'] - factors['idio_vol'] -
            factors['debt_ratio'] + factors['gross_margin'] + factors['wc_turnover'] + factors['cash_flow_to_assets'] -
            factors['turnover_rate']
        )
        # factors['score'] = (
        #     -factors['pe'] - factors['pb'] +
        #     factors['roe'] + factors['revenue_growth'] -
        #     factors['debt_ratio'] +
        #     0.5 * factors['short_momentum'] + 0.5 * factors['reversal'] - factors['rsi'] +
        #     factors['low_vol']
        # )
        # 波动率过滤
        factors['score'] = factors['score'].where(factors['raw_volatility'] < factors['raw_volatility'].quantile(0.75), np.nan)
        factors['symbol'] = symbol

        return factors
    except Exception as e:
        logging.error(f"股票 {symbol} 因子计算失败: {str(e)}")
        return pd.DataFrame()

# 检查因子缓存
all_factors_file = "data_clean/all_factors.csv"
if os.path.exists(all_factors_file) and os.path.getsize(all_factors_file) > 0 :
    logging.info("Loading cached all_factors.csv")
    all_factors = pd.read_csv(all_factors_file, parse_dates=['date'], index_col='date')
else:
    logging.info("Calculating factors in parallel")
    all_factors_list = Parallel(n_jobs=4, verbose=5)(delayed(calculate_factors)(symbol) for symbol in symbols)
    all_factors_list = [f for f in all_factors_list if not f.empty]
    if not all_factors_list:
        raise ValueError("所有股票因子数据为空")
    all_factors = pd.concat(all_factors_list, axis=0)
    all_factors.to_csv(all_factors_file, encoding='utf-8')

# 选股函数
def select_stocks(factors_dict, rebalance_date, index_weights, alpha=0.5):
    try:
        rebalance_date = pd.Timestamp(rebalance_date)
        factors = factors_dict.get(rebalance_date)
        if factors is None or factors.empty:
            logging.warning(f"因子数据为空，日期: {rebalance_date}")
            return {}
        
        factors = factors.sort_values('score', ascending=False)
        num_stocks = min(250, len(factors))
        selected = factors['symbol'].iloc[:num_stocks].tolist()
        
        index_weights = index_weights[index_weights['日期'] <= rebalance_date].tail(1)
        index_weights_dict = {row['stock_code']: row['权重']/100 for _, row in index_weights.iterrows()}
        
        scores = factors.loc[factors['symbol'].isin(selected), 'score']
        factor_weights = {stock: score / scores.sum() for stock, score in zip(selected, scores)}
        return factor_weights
    except Exception as e:
        logging.error(f"选股失败，日期: {rebalance_date}, 错误: {str(e)}")
        return {}

# 指数增强策略
class IndexEnhanceStrategy(bt.Strategy):
    params = (
        ('alpha', 0.5),
        ('stop_loss_stock', 0.10),
        ('stop_loss_portfolio', 0.20),
    )

    def __init__(self):
        self.all_factors = pd.read_csv("data_clean/all_factors.csv", parse_dates=['date'], index_col='date')
        self.factors_dict = {date: df for date, df in self.all_factors.groupby(self.all_factors.index)}
        self.index_weights = pd.read_csv("data_clean/index/csi500_weights.csv", parse_dates=['日期'])
        self.weights = {}
        self.prev_weights = {}
        self.portfolio_values = []
        self.dates = []
        self.initial_value = self.broker.getvalue()

    def next(self):
        dt = pd.Timestamp(self.datas[0].datetime.date(0))
        portfolio_value = self.broker.getvalue()
        self.dates.append(dt)
        self.portfolio_values.append(portfolio_value)

        if pd.isna(self.datas[0].close[0]) or self.datas[0].close[0] <= 0:
            return

        if (self.initial_value - portfolio_value) / self.initial_value > self.params.stop_loss_portfolio:
            for data in self.datas[1:]:
                self.close(data=data)
            return

        for data in self.datas[1:]:
            position = self.getposition(data)
            if position.size > 0:
                entry_price = position.price
                current_price = data.close[0]
                if pd.isna(current_price) or current_price <= 0:
                    continue
                if (entry_price - current_price) / entry_price > self.params.stop_loss_stock:
                    self.close(data=data)

        new_weights = select_stocks(self.factors_dict, dt, self.index_weights, self.params.alpha)
        if not new_weights:
            return

        self.weights = new_weights
        self.prev_weights = new_weights.copy()

        for data in self.datas[1:]:
            symbol = data._name
            target_weight = self.weights.get(symbol, 0)
            current_pos = self.getposition(data).size
            price = data.close[0]
            if target_weight > 0:
                if pd.isna(price) or price <= 0:
                    continue
                try:
                    target_shares = int(target_weight * portfolio_value / price / 100) * 100
                    if target_shares > current_pos:
                        self.buy(data=data, size=target_shares - current_pos)
                    elif target_shares < current_pos:
                        self.sell(data=data, size=current_pos - target_shares)
                except (ValueError, TypeError):
                    continue

# 回测
cerebro = bt.Cerebro(stdstats=False)
index_data = pd.read_csv("data_clean/index/csi500_index.csv", parse_dates=['date'], index_col='date')
if index_data.empty or 'close' not in index_data.columns or index_data['close'].isna().any() or (index_data['close'] <= 0).any():
    raise ValueError("Invalid index data")
index_data = index_data.sort_index()
cerebro.adddata(bt.feeds.PandasData(
    dataname=index_data,
    name="csi500",
    datetime=None,
    open='open',
    high='high',
    low='low',
    close='close',
    volume='volume',
    fromdate=pd.to_datetime("2020-01-04"),
    todate=pd.to_datetime("2022-12-30")
))

data_feed_count = 1
hdf_file = "data_clean/daily_cleaned.h5"
with pd.HDFStore(hdf_file, mode='r') as store:
    for symbol in symbols:
        try:
            data = store[f'stock_{symbol}']
            if data.empty or 'close' not in data.columns or data['close'].isna().any() or (data['close'] <= 0).any():
                continue
            data = data.sort_index()
            data = data.reindex(index_data.index).ffill().bfill()
            data['open'] = data['open'].fillna(data['close'])
            data['high'] = data['high'].fillna(data['close'])
            data['low'] = data['low'].fillna(data['close'])
            data['volume'] = data['volume'].fillna(0)
            if data['close'].isna().any() or (data['close'] <= 0).any():
                continue
            cerebro.adddata(bt.feeds.PandasData(
                dataname=data,
                name=symbol,
                datetime=None,
                open='open',
                high='high',
                low='low',
                close='close',
                volume='volume',
                fromdate=pd.to_datetime("2020-01-04"),
                todate=pd.to_datetime("2022-12-30")
            ))
            data_feed_count += 1
        except Exception:
            continue

if data_feed_count <= 1:
    raise ValueError("No valid stock data feeds loaded")

cerebro.addstrategy(IndexEnhanceStrategy)
cerebro.broker.setcash(1000000.0)
cerebro.broker.setcommission(commission=0.001)
cerebro.addsizer(bt.sizers.FixedSize, stake=100)
cerebro.addanalyzer(bt.analyzers.Returns, _name='returns')
cerebro.addanalyzer(bt.analyzers.TradeAnalyzer, _name='trades')
cerebro.addanalyzer(bt.analyzers.DrawDown, _name='drawdown')

results = cerebro.run(runonce=True, preload=True, exactbars=0)
strat = results[0]

# 处理回报数据
portfolio_values = pd.Series(strat.portfolio_values, index=pd.to_datetime(strat.dates))
portfolio_daily_returns = portfolio_values.pct_change().fillna(0)
portfolio_cum_returns = (1 + portfolio_daily_returns).cumprod() - 1
index_daily_returns = index_data['close'].pct_change().fillna(0)
index_cum_returns = (1 + index_daily_returns).cumprod() - 1

# 对齐日期
common_index = portfolio_daily_returns.index.intersection(index_daily_returns.index)
portfolio_daily_returns = portfolio_daily_returns.loc[common_index]
portfolio_cum_returns = portfolio_cum_returns.loc[common_index]
index_daily_returns = index_daily_returns.loc[common_index]
index_cum_returns = index_cum_returns.loc[common_index]

# 生成 Excel 表格
returns_df = pd.DataFrame({
    'Date': common_index,
    'Portfolio_Daily_Return': portfolio_daily_returns,
    'CSI500_Daily_Return': index_daily_returns
})
excel_file = "data_clean/daily_returns_comparison.xlsx"
returns_df.to_excel(excel_file, index=False, sheet_name='Daily Returns')

# 格式化 Excel
wb = openpyxl.load_workbook(excel_file)
ws = wb['Daily Returns']
ws['A1'] = '日期'
ws['B1'] = '组合日收益率'
ws['C1'] = '中证500日收益率'

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.number_format = 'YYYY-MM-DD'

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
    for cell in row:
        cell.number_format = '0.00%'

wb.save(excel_file)

# 生成累计收益率图表
fig = go.Figure()
if not portfolio_cum_returns.empty and portfolio_cum_returns.isna().sum() < len(portfolio_cum_returns):
    fig.add_trace(
        go.Scatter(
            x=portfolio_cum_returns.index,
            y=portfolio_cum_returns,
            mode='lines',
            name='Portfolio Cumulative Return',
            line=dict(color='blue', width=2),
            connectgaps=True
        )
    )
    fig.add_annotation(
        x=portfolio_cum_returns.index[-1],
        y=portfolio_cum_returns.iloc[-1],
        text=f"Portfolio: {portfolio_cum_returns.iloc[-1]*100:.2f}%",
        showarrow=True,
        arrowhead=2,
        ax=20,
        ay=-30,
        font=dict(color='blue')
    )

fig.add_trace(
    go.Scatter(
        x=index_cum_returns.index,
        y=index_cum_returns,
        mode='lines',
        name='CSI 500 Cumulative Return',
        line=dict(color='red', dash='dash', width=2),
        connectgaps=True
    )
)
fig.add_annotation(
    x=index_cum_returns.index[-1],
    y=index_cum_returns.iloc[-1],
    text=f"CSI 500: {index_cum_returns.iloc[-1]*100:.2f}%",
    showarrow=True,
    arrowhead=2,
    ax=20,
    ay=30,
    font=dict(color='red')
)

fig.update_layout(
    title="指数增强策略累计收益率 (2020-01-04 至 2022-12-30)",
    height=500,
    showlegend=True,
    legend_title_text="Portfolio vs. CSI 500",
    xaxis=dict(
        title="日期",
        type='date',
        tickformat="%Y-%m-%d",
        range=[pd.to_datetime("2020-01-04"), pd.to_datetime("2022-12-30")]
    ),
    yaxis=dict(title="累计收益率 (%)", tickformat=".2%"),
    hovermode="x unified"
)

fig.write_html("data_clean/cumulative_returns.html")

# 性能指标
total_return = strat.analyzers.returns.get_analysis()['rtot']
days = (portfolio_values.index[-1] - portfolio_values.index[0]).days
annualized_return = (1 + total_return) ** (252 / days) - 1
num_trades = strat.analyzers.trades.get_analysis()['total']['total']
max_drawdown = strat.analyzers.drawdown.get_analysis()['max']['drawdown'] / 100

print(f"最终资金: {cerebro.broker.getvalue():.2f}")
print(f"总回报: {total_return*100:.2f}%")
print(f"年化回报: {annualized_return*100:.2f}%")
print(f"最大回撤: {-max_drawdown*100:.2f}%")
print(f"总交易次数: {num_trades}")
print(f"日收益率比较已导出至: {excel_file}")
print(f"累计收益率图表已生成: data_clean/cumulative_returns.html")