import akshare as ak
import pandas as pd
import backtrader as bt
import logging
import numpy as np
import talib
import os
import time
from tqdm import tqdm
from datetime import datetime, timedelta
from joblib import Parallel, delayed
import openpyxl
from openpyxl.styles import numbers
import plotly.graph_objects as go
import statsmodels.api as sm

# 设置日志
logger = logging.getLogger()
logger.setLevel(logging.INFO)

log_file = 'strategy.log'
if os.path.exists(log_file):
    os.remove(log_file)

file_handler = logging.FileHandler(log_file, encoding='utf-8')
file_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# 创建目录
os.makedirs("data/daily/cleaned", exist_ok=True)
os.makedirs("data/pe_pb/cleaned", exist_ok=True)
os.makedirs("data/financial/cleaned", exist_ok=True)
os.makedirs("data/dividend/cleaned", exist_ok=True)

# 数据层：获取中证500成分股和数据
def get_stock_code(symbol, exchange):
    """将6位股票代码转换为带交易所前缀的格式"""
    if exchange == "上海证券交易所":
        return f"sh{symbol}"
    elif exchange == "深圳证券交易所":
        return f"sz{symbol}"
    return symbol

def getData():
    start_date = "20200104"
    end_date = "20221230"
    
    # 检查缓存
    cache_file = "data/csi500_data_cache.pkl"
    if os.path.exists(cache_file):
        cached_data = pd.read_pickle(cache_file)
        return cached_data['symbols'], cached_data['index_weights']
    
    # 获取中证500成分股
    csi500_constituents = ak.index_stock_cons_weight_csindex(symbol="000905")
    if csi500_constituents is None or csi500_constituents.empty:
        logging.error("CSI 500 constituents data is empty or None")
        raise ValueError("CSI 500 constituents data is empty or None")
    csi500_constituents['stock_code'] = csi500_constituents.apply(
        lambda x: get_stock_code(x['成分券代码'], x['交易所']), axis=1
    )
    csi500_constituents['日期'] = pd.to_datetime(csi500_constituents['日期'])
    csi500_constituents.to_csv("data/csi500_weights.csv", encoding='utf-8')
    time.sleep(2)
    
    symbols = csi500_constituents['stock_code'].unique().tolist()[:20]
    
    # 获取指数数据
    index_data = ak.index_zh_a_hist(symbol="000905", period="daily", start_date=start_date, end_date=end_date)
    if index_data is None or index_data.empty:
        logging.error("Index data is empty or None")
        raise ValueError("Index data is empty or None")
    index_data['date'] = pd.to_datetime(index_data['日期'])
    index_data = index_data.rename(columns={
        '开盘': 'open', '收盘': 'close', '最高': 'high', '最低': 'low', '成交量': 'volume'
    })
    index_data.set_index('date', inplace=True)
    index_data['open'] = index_data['open'].ffill().bfill()
    index_data['high'] = index_data['high'].ffill().bfill()
    index_data['low'] = index_data['low'].ffill().bfill()
    index_data['close'] = index_data['close'].ffill().bfill()
    index_data['volume'] = index_data['volume'].replace([np.inf, -np.inf], 0).fillna(0)
    index_data = index_data.sort_index()
    if index_data['close'].isna().any() or (index_data['close'] <= 0).any():
        logging.error(f"Index data contains invalid close prices")
        raise ValueError("Index data contains invalid close prices")
    index_data.index.name = 'date'
    index_data.to_csv("data/csi500_index.csv", encoding='utf-8')
    time.sleep(2)

    # 合并日K线数据到 HDF5
    hdf_file = "data/daily_cleaned.h5"
    valid_symbols = []
    with pd.HDFStore(hdf_file, mode='w') as store:
        for symbol in tqdm(symbols, desc="Processing stocks", unit="stock"):
            cleaned_daily_file = f"data/daily/cleaned/stock_{symbol}_cleaned.csv"
            cleaned_pe_pb_file = f"data/pe_pb/cleaned/stock_{symbol}_pe_pb_cleaned.csv"
            cleaned_financial_file = f"data/financial/cleaned/stock_{symbol}_financial_cleaned.csv"
            cleaned_dividend_file = f"data/dividend/cleaned/stock_{symbol}_dividend_cleaned.csv"

            if (os.path.exists(cleaned_daily_file) and os.path.getsize(cleaned_daily_file) > 0 and
                os.path.exists(cleaned_pe_pb_file) and os.path.getsize(cleaned_pe_pb_file) > 0 and
                os.path.exists(cleaned_financial_file) and os.path.getsize(cleaned_financial_file) > 0 and
                os.path.exists(cleaned_dividend_file) and os.path.getsize(cleaned_dividend_file) > 0):
                try:
                    stock_data = pd.read_csv(cleaned_daily_file, parse_dates=['date'], index_col='date')
                    if not (stock_data['close'].isna().any() or (stock_data['close'] <= 0).any() or stock_data['close'].lt(0).any()):
                        store[f'stock_{symbol}'] = stock_data
                        valid_symbols.append(symbol)
                        continue
                except Exception as e:
                    logging.warning(f"股票 {symbol} 读取现有CSV失败: {str(e)}")

            symbol_code = symbol[2:]
            try:
                time.sleep(1)
                stock_info = ak.stock_individual_info_em(symbol=symbol_code)
                if stock_info is None or stock_info.empty:
                    continue

                # 日K线数据
                time.sleep(1)
                stock_data = ak.stock_zh_a_daily(symbol=symbol, adjust="qfq", start_date=start_date, end_date=end_date)
                if stock_data is None or stock_data.empty:
                    continue
                stock_data['date'] = pd.to_datetime(stock_data['date'])
                stock_data = stock_data.rename(columns={
                    'open': 'open', 'high': 'high', 'low': 'low', 'close': 'close', 'volume': 'volume'
                })
                stock_data.set_index('date', inplace=True)
                stock_data['volume'] = stock_data['volume'].replace([np.inf, -np.inf], 0).fillna(0)
                
                stock_data['close'] = stock_data['close'].ffill().bfill()
                if stock_data['close'].isna().any() or (stock_data['close'] <= 0).any() or stock_data['close'].lt(0).any():
                    mean_close = stock_data['close'].mean()
                    if pd.isna(mean_close) or mean_close <= 0:
                        continue
                    stock_data['close'] = stock_data['close'].fillna(mean_close)
                    if stock_data['close'].lt(0).any():
                        continue

                # PE/PB 数据
                time.sleep(1)
                pe_pb_data = ak.stock_a_indicator_lg(symbol=symbol_code)
                if pe_pb_data is None or pe_pb_data.empty:
                    continue
                pe_pb_data['date'] = pd.to_datetime(pe_pb_data['trade_date'])
                pe_pb_data.set_index('date', inplace=True)
                pe_pb_data = pe_pb_data[(pe_pb_data.index >= pd.to_datetime(start_date)) & (pe_pb_data.index <= pd.to_datetime(end_date))]
                
                # 财务数据
                time.sleep(1)
                financial_data = ak.stock_financial_analysis_indicator(symbol=symbol_code, start_year="2020")
                if financial_data is None or financial_data.empty:
                    continue
                financial_data['date'] = pd.to_datetime(financial_data['日期'])
                financial_data.set_index('date', inplace=True)
                financial_data = financial_data[(financial_data.index >= pd.to_datetime(start_date)) & (financial_data.index <= pd.to_datetime(end_date))]
                
                # 股息率数据
                try:
                    time.sleep(1)
                    dividend_data = ak.stock_dividend_cninfo(symbol=symbol_code)
                    if dividend_data is None or dividend_data.empty:
                        dividend_data = pd.DataFrame({'股权登记日': [], '分红类型': [], '除权日': []})
                    dividend_data['date'] = pd.to_datetime(dividend_data['股权登记日'])
                    dividend_data = dividend_data[dividend_data['date'].notna()]
                    dividend_data.set_index('date', inplace=True)
                    dividend_data = dividend_data[(dividend_data.index >= pd.to_datetime(start_date)) & (dividend_data.index <= pd.to_datetime(end_date))]

                    # 计算每股股息（解析分红方案，如“10 派 2 元”）
                    def parse_dividend_scheme(scheme):
                        if pd.isna(scheme) or not isinstance(scheme, str):
                            return 0
                        match = re.search(r'10\s*派\s*(\d+\.?\d*)', scheme)
                        if match:
                            cash = float(match.group(1))
                            return cash / 10  # 每股股息（10 股派 X 元）
                        return 0

                    dividend_data['per_share_dividend'] = dividend_data['派息比例']/10
                    
                    # 计算年化股息（假设每年分红一次）
                    dividend_data = dividend_data.groupby(dividend_data.index.to_period('Y')).agg({
                        'per_share_dividend': 'sum'
                    }).reset_index()
                    dividend_data['date'] = dividend_data['index'].apply(lambda x: x.to_timestamp(how='end'))
                    dividend_data.set_index('date', inplace=True)
                    dividend_data.drop(columns=['index'], inplace=True)
                    
                    # 匹配股价计算股息率
                    dividend_data = dividend_data.join(stock_data['close'], how='left')
                    dividend_data['dividend_yield'] = dividend_data['per_share_dividend'] / dividend_data['close'].replace(0, np.nan)
                    dividend_data['dividend_yield'] = dividend_data['dividend_yield'].replace([np.inf, -np.inf], 0).clip(lower=0)
                except Exception as e:
                    logging.warning(f"股票 {symbol} 股息率数据获取失败: {str(e)}")
                    dividend_data = pd.DataFrame(index=pd.date_range(start=start_date, end=end_date, freq='D'), columns=['dividend_yield'])
                    dividend_data['dividend_yield'] = 0
                
                # 经营现金流数据
                cash_flow_data = ak.stock_financial_report_sina(stock=symbol_code, symbol="cash_flow")
                if cash_flow_data is None or cash_flow_data.empty:
                    cash_flow_data = pd.DataFrame({'报告期': [], '经营活动产生的现金流量净额': [], '资产总计': []})
                cash_flow_data['date'] = pd.to_datetime(cash_flow_data['报告期'])
                cash_flow_data.set_index('date', inplace=True)
                cash_flow_data = cash_flow_data[(cash_flow_data.index >= pd.to_datetime(start_date)) & (cash_flow_data.index <= pd.to_datetime(end_date))]
                # 合并财务和现金流数据

                # 确保所需字段
                required_cols = ['净资产收益率(%)', '主营业务收入增长率(%)', '资产负债率(%)',
                                '净利润同比增长(%)', '总资产增长率(%)', '销售毛利率(%)',
                                '流动资产周转率(次)', '资产的经营现金流量回报率(%)']
                for col in required_cols:
                    if col not in financial_data.columns:
                        financial_data[col] = 0
                
                base_index = pd.date_range(start=start_date, end=end_date, freq='D')
                stock_data = stock_data.reindex(base_index).ffill().bfill()
                pe_pb_data = pe_pb_data.reindex(base_index).ffill().bfill()
                financial_data = financial_data.reindex(base_index).ffill().bfill()
                dividend_data = dividend_data.reindex(base_index).ffill().bfill()

                stock_data['volume'] = stock_data['volume'].fillna(0)
                stock_data['close'] = stock_data['close'].fillna(stock_data['close'].mean())
                stock_data['open'] = stock_data['open'].fillna(stock_data['close'])
                stock_data['high'] = stock_data['high'].fillna(stock_data['close'])
                stock_data['low'] = stock_data['low'].fillna(stock_data['close'])
                pe_pb_data['pe'] = pe_pb_data['pe'].fillna(pe_pb_data['pe'].mean()).fillna(20)
                pe_pb_data['pb'] = pe_pb_data['pb'].fillna(pe_pb_data['pb'].mean()).fillna(2)
                financial_data['净资产收益率(%)'] = financial_data['净资产收益率(%)'].fillna(financial_data['净资产收益率(%)'].mean()).fillna(0)
                financial_data['主营业务收入增长率(%)'] = financial_data['主营业务收入增长率(%)'].fillna(financial_data['主营业务收入增长率(%)'].mean()).fillna(0)
                financial_data['资产负债率(%)'] = financial_data['资产负债率(%)'].fillna(financial_data['资产负债率(%)'].mean()).fillna(0)
                financial_data['净利润同比增长(%)'] = financial_data['净利润同比增长(%)'].fillna(financial_data['净利润同比增长(%)'].mean()).fillna(0)
                financial_data['总资产增长率(%)'] = financial_data['总资产增长率(%)'].fillna(financial_data['总资产增长率(%)'].mean()).fillna(0)
                financial_data['销售毛利率(%)'] = financial_data['销售毛利率(%)'].fillna(financial_data['销售毛利率(%)'].mean()).fillna(0)
                financial_data['流动资产周转率(次)'] = financial_data['流动资产周转率(次)'].fillna(financial_data['流动资产周转率(次)'].mean()).fillna(0)
                financial_data['资产的经营现金流量回报率(%)'] = financial_data['资产的经营现金流量回报率(%)'].fillna(financial_data['资产的经营现金流量回报率(%)'].mean()).fillna(0)
                dividend_data['dividend_yield'] = dividend_data['dividend_yield'].fillna(0)

                if (stock_data['close'].isna().any() or (stock_data['close'] <= 0).any() or stock_data['close'].lt(0).any() or
                    pe_pb_data['pe'].isna().any() or pe_pb_data['pb'].isna().any() or
                    financial_data['净资产收益率(%)'].isna().any() or
                    financial_data['主营业务收入增长率(%)'].isna().any() or
                    financial_data['资产负债率(%)'].isna().any() or
                    financial_data['净利润同比增长(%)'].isna().any() or
                    financial_data['总资产增长率(%)'].isna().any() or
                    financial_data['销售毛利率(%)'].isna().any() or
                    financial_data['流动资产周转率(次)'].isna().any() or
                    financial_data['资产的经营现金流量回报率(%)'].isna().any() or
                    financial_data['dividend_yield'].isna().any() ):
                    continue

                stock_data.index.name = 'date'
                pe_pb_data.index.name = 'date'
                financial_data.index.name = 'date'
                dividend_data.index.name = 'date'

                stock_data.to_csv(cleaned_daily_file, encoding='utf-8')
                pe_pb_data.to_csv(cleaned_pe_pb_file, encoding='utf-8')
                financial_data.to_csv(cleaned_financial_file, encoding='utf-8')
                dividend_data.to_csv(cleaned_dividend_file, encoding='utf-8')
                store[f'stock_{symbol}'] = stock_data
                valid_symbols.append(symbol)
            except Exception as e:
                logging.error(f"股票 {symbol} 数据处理失败: {str(e)}")
                continue
    
    pd.to_pickle({'symbols': valid_symbols, 'index_weights': csi500_constituents}, cache_file)
    return valid_symbols, csi500_constituents

symbols, index_weights = getData()

# 因子层：计算多因子
column_mapping = {
    "净资产收益率(%)": "roe",
    "主营业务收入增长率(%)": "revenue_growth",
    "资产负债率(%)": "debt_ratio",
    "净利润同比增长(%)": "net_profit_growth",
    "总资产增长率(%)": "asset_growth",
    "销售毛利率(%)": "gross_margin",
    "流动资产周转率(次)": "wc_turnover",
}

def calculate_factors(symbol):
    try:
        # 加载数据
        stock_data = pd.read_csv(f"data/daily/cleaned/stock_{symbol}_cleaned.csv", parse_dates=['date'], index_col='date')
        pe_pb_data = pd.read_csv(f"data/pe_pb/cleaned/stock_{symbol}_pe_pb_cleaned.csv", parse_dates=['date'], index_col='date')
        financial_data = pd.read_csv(f"data/financial/cleaned/stock_{symbol}_financial_cleaned.csv", parse_dates=['date'], index_col='date')
        dividend_data = pd.read_csv(f"data/dividend/cleaned/stock_{symbol}_dividend_cleaned.csv", parse_dates=['date'], index_col='date')
        index_data = pd.read_csv("data/csi500_index.csv", parse_dates=['date'], index_col='date')

        # 数据完整性检查
        if (stock_data['close'].isna().all() or (stock_data['close'] <= 0).all() or stock_data['close'].lt(0).any() or
            'pe' not in pe_pb_data.columns or pe_pb_data['pe'].isna().all() or
            'pb' not in pe_pb_data.columns or pe_pb_data['pb'].isna().all() or
            '净资产收益率(%)' not in financial_data.columns or financial_data['净资产收益率(%)'].isna().all() or
            'dividend_yield' not in dividend_data.columns or dividend_data['dividend_yield'].isna().all()):
            logging.warning(f"股票 {symbol} 数据不完整")
            return pd.DataFrame()

        factors = pd.DataFrame(index=stock_data.index)
        financial_data = financial_data.rename(columns=column_mapping)

        # 1. 价值因子
        factors['pe'] = pe_pb_data['pe'].ffill().bfill().fillna(pe_pb_data['pe'].mean()).fillna(20)
        factors['pb'] = pe_pb_data['pb'].ffill().bfill().fillna(pe_pb_data['pb'].mean()).fillna(2)
        factors['ps'] = financial_data.get('总收入', stock_data['close'] * 100).div(stock_data['close']).ffill().bfill().fillna(20)
        factors['dividend_yield'] = dividend_data['dividend_yield'].ffill().bfill().fillna(0)

        # 2. 成长因子
        factors['roe'] = financial_data['roe'].ffill().bfill().fillna(financial_data['roe'].mean()).fillna(0)
        factors['revenue_growth'] = financial_data['revenue_growth'].ffill().bfill().fillna(financial_data['revenue_growth'].mean()).fillna(0)
        factors['net_profit_growth'] = financial_data.get('net_profit_growth', 0).ffill().bfill().fillna(0)
        factors['asset_growth'] = financial_data.get('asset_growth', 0).ffill().bfill().fillna(0)

        # 3. 动量因子
        factors['short_momentum'] = stock_data['close'].pct_change(20).fillna(0)
        factors['long_momentum'] = stock_data['close'].pct_change(120).fillna(0)
        factors['reversal'] = -stock_data['close'].pct_change(5).fillna(0)

        # 4. 技术因子
        factors['rsi'] = pd.Series(talib.RSI(stock_data['close'], timeperiod=14), index=stock_data.index).fillna(50)
        upper, middle, lower = talib.BBANDS(stock_data['close'], timeperiod=20)
        factors['bb_width'] = ((upper - lower) / middle).fillna(0)
        macd, signal, _ = talib.MACD(stock_data['close'], fastperiod=12, slowperiod=26, signalperiod=9)
        factors['macd_diff'] = (pd.Series(macd, index=stock_data.index) - pd.Series(signal, index=stock_data.index)).fillna(0)

        # 5. 波动率因子
        factors['raw_volatility'] = stock_data['close'].pct_change().rolling(20, min_periods=10).std().fillna(0.1)
        if (factors['raw_volatility'] < 0).any():
            logging.error(f"股票 {symbol} 包含负原始波动率")
            return pd.DataFrame()
        factors['low_vol'] = -factors['raw_volatility']
        # 特异波动率
        market_returns = index_data['close'].pct_change().fillna(0)
        stock_returns = stock_data['close'].pct_change().fillna(0)
        idio_vol = []
        for t in stock_data.index:
            window = pd.DataFrame({
                'stock': stock_returns.loc[:t].tail(60),
                'market': market_returns.loc[:t].tail(60)
            }).dropna()
            if len(window) < 20:
                idio_vol.append(0)
                continue
            X = sm.add_constant(window['market'])
            model = sm.OLS(window['stock'], X).fit()
            residuals = model.resid
            idio_vol.append(residuals.std())
        factors['idio_vol'] = pd.Series(idio_vol, index=stock_data.index).fillna(0.1)

        # 6. 质量因子
        factors['debt_ratio'] = financial_data['debt_ratio'].ffill().bfill().fillna(financial_data['debt_ratio'].mean()).fillna(50)
        factors['gross_margin'] = financial_data.get('gross_margin', 0).ffill().bfill().fillna(0)
        factors['wc_turnover'] = financial_data['wc_turnover'].ffill().bfill().fillna(financial_data['wc_turnover'].mean()).fillna(0)
        factors['cash_flow_to_assets'] = financial_data['cash_flow_to_assets'].ffill().bfill().fillna(financial_data['cash_flow_to_assets'].mean()).fillna(0)

        # 7. 市场因子
        factors['turnover_rate'] = (stock_data['volume'] / stock_data['volume'].mean()).rolling(20).mean().fillna(0)

        # 因子列表
        factor_columns = [
            'pe', 'pb', 'ps', 'dividend_yield',
            'roe', 'revenue_growth', 'net_profit_growth', 'asset_growth',
            'short_momentum', 'long_momentum', 'reversal',
            'rsi', 'bb_width', 'macd_diff',
            'low_vol', 'idio_vol',
            'debt_ratio', 'gross_margin', 'wc_turnover', 'cash_flow_to_assets',
            'turnover_rate'
        ]

        # 标准化因子
        for col in factor_columns:
            mean = factors[col].mean()
            std = factors[col].std()
            factors[col] = factors[col] - mean if pd.isna(std) or std == 0 else (factors[col] - mean) / std

        # 等权计算 score
        factors['score'] = (
            -factors['pe'] - factors['pb'] - factors['ps'] + factors['dividend_yield'] +
            factors['roe'] + factors['revenue_growth'] + factors['net_profit_growth'] + factors['asset_growth'] +
            factors['short_momentum'] + factors['long_momentum'] + factors['reversal'] -
            factors['rsi'] - factors['bb_width'] + factors['macd_diff'] +
            factors['low_vol'] - factors['idio_vol'] -
            factors['debt_ratio'] + factors['gross_margin'] + factors['wc_turnover'] + factors['cash_flow_to_assets'] -
            factors['turnover_rate']
        )

        # 波动率过滤
        factors['score'] = factors['score'].where(factors['raw_volatility'] < factors['raw_volatility'].quantile(0.75), np.nan)
        factors['symbol'] = symbol

        return factors
    except Exception as e:
        logging.error(f"股票 {symbol} 因子计算失败: {str(e)}")
        return pd.DataFrame()

# 检查因子缓存
all_factors_file = "data/all_factors.csv"
if os.path.exists(all_factors_file) and os.path.getsize(all_factors_file) > 0:
    logging.info("Loading cached all_factors.csv")
    all_factors = pd.read_csv(all_factors_file, parse_dates=['date'], index_col='date')
else:
    logging.info("Calculating factors in parallel")
    all_factors_list = Parallel(n_jobs=4, verbose=5)(delayed(calculate_factors)(symbol) for symbol in symbols)
    all_factors_list = [f for f in all_factors_list if not f.empty]
    if not all_factors_list:
        raise ValueError("所有股票因子数据为空")
    all_factors = pd.concat(all_factors_list, axis=0)
    all_factors.to_csv(all_factors_file, encoding='utf-8')

# 选股函数
def select_stocks(factors_dict, rebalance_date, index_weights, alpha=0.5):
    try:
        rebalance_date = pd.Timestamp(rebalance_date)
        factors = factors_dict.get(rebalance_date)
        if factors is None or factors.empty:
            logging.warning(f"因子数据为空，日期: {rebalance_date}")
            return {}
        
        factors = factors.sort_values('score', ascending=False)
        num_stocks = min(250, len(factors))
        selected = factors['symbol'].iloc[:num_stocks].tolist()
        
        index_weights = index_weights[index_weights['日期'] <= rebalance_date].tail(1)
        index_weights_dict = {row['stock_code']: row['权重']/100 for _, row in index_weights.iterrows()}
        
        scores = factors.loc[factors['symbol'].isin(selected), 'score']
        factor_weights = {stock: score / scores.sum() for stock, score in zip(selected, scores)}
        return factor_weights
    except Exception as e:
        logging.error(f"选股失败，日期: {rebalance_date}, 错误: {str(e)}")
        return {}

# 指数增强策略
class IndexEnhanceStrategy(bt.Strategy):
    params = (
        ('alpha', 0.5),
        ('stop_loss_stock', 0.10),
        ('stop_loss_portfolio', 0.20),
    )

    def __init__(self):
        self.all_factors = pd.read_csv("data/all_factors.csv", parse_dates=['date'], index_col='date')
        self.factors_dict = {date: df for date, df in self.all_factors.groupby(self.all_factors.index)}
        self.index_weights = pd.read_csv("data/csi500_weights.csv", parse_dates=['日期'])
        self.weights = {}
        self.prev_weights = {}
        self.portfolio_values = []
        self.dates = []
        self.initial_value = self.broker.getvalue()

    def next(self):
        dt = pd.Timestamp(self.datas[0].datetime.date(0))
        portfolio_value = self.broker.getvalue()
        self.dates.append(dt)
        self.portfolio_values.append(portfolio_value)

        if pd.isna(self.datas[0].close[0]) or self.datas[0].close[0] <= 0:
            return

        if (self.initial_value - portfolio_value) / self.initial_value > self.params.stop_loss_portfolio:
            for data in self.datas[1:]:
                self.close(data=data)
            return

        for data in self.datas[1:]:
            position = self.getposition(data)
            if position.size > 0:
                entry_price = position.price
                current_price = data.close[0]
                if pd.isna(current_price) or current_price <= 0:
                    continue
                if (entry_price - current_price) / entry_price > self.params.stop_loss_stock:
                    self.close(data=data)

        new_weights = select_stocks(self.factors_dict, dt, self.index_weights, self.params.alpha)
        if not new_weights:
            return

        self.weights = new_weights
        self.prev_weights = new_weights.copy()

        for data in self.datas[1:]:
            symbol = data._name
            target_weight = self.weights.get(symbol, 0)
            current_pos = self.getposition(data).size
            price = data.close[0]
            if target_weight > 0:
                if pd.isna(price) or price <= 0:
                    continue
                try:
                    target_shares = int(target_weight * portfolio_value / price / 100) * 100
                    if target_shares > current_pos:
                        self.buy(data=data, size=target_shares - current_pos)
                    elif target_shares < current_pos:
                        self.sell(data=data, size=current_pos - target_shares)
                except (ValueError, TypeError):
                    continue

# 回测
cerebro = bt.Cerebro(stdstats=False)
index_data = pd.read_csv("data/csi500_index.csv", parse_dates=['date'], index_col='date')
if index_data.empty or 'close' not in index_data.columns or index_data['close'].isna().any() or (index_data['close'] <= 0).any():
    raise ValueError("Invalid index data")
index_data = index_data.sort_index()
cerebro.adddata(bt.feeds.PandasData(
    dataname=index_data,
    name="csi500",
    datetime=None,
    open='open',
    high='high',
    low='low',
    close='close',
    volume='volume',
    fromdate=pd.to_datetime("2020-01-04"),
    todate=pd.to_datetime("2022-12-30")
))

data_feed_count = 1
hdf_file = "data/daily_cleaned.h5"
with pd.HDFStore(hdf_file, mode='r') as store:
    for symbol in symbols:
        try:
            data = store[f'stock_{symbol}']
            if data.empty or 'close' not in data.columns or data['close'].isna().any() or (data['close'] <= 0).any():
                continue
            data = data.sort_index()
            data = data.reindex(index_data.index).ffill().bfill()
            data['open'] = data['open'].fillna(data['close'])
            data['high'] = data['high'].fillna(data['close'])
            data['low'] = data['low'].fillna(data['close'])
            data['volume'] = data['volume'].fillna(0)
            if data['close'].isna().any() or (data['close'] <= 0).any():
                continue
            cerebro.adddata(bt.feeds.PandasData(
                dataname=data,
                name=symbol,
                datetime=None,
                open='open',
                high='high',
                low='low',
                close='close',
                volume='volume',
                fromdate=pd.to_datetime("2020-01-04"),
                todate=pd.to_datetime("2022-12-30")
            ))
            data_feed_count += 1
        except Exception:
            continue

if data_feed_count <= 1:
    raise ValueError("No valid stock data feeds loaded")

cerebro.addstrategy(IndexEnhanceStrategy)
cerebro.broker.setcash(1000000.0)
cerebro.broker.setcommission(commission=0.001)
cerebro.addsizer(bt.sizers.FixedSize, stake=100)
cerebro.addanalyzer(bt.analyzers.Returns, _name='returns')
cerebro.addanalyzer(bt.analyzers.TradeAnalyzer, _name='trades')
cerebro.addanalyzer(bt.analyzers.DrawDown, _name='drawdown')

results = cerebro.run(runonce=True, preload=True, exactbars=0)
strat = results[0]

# 处理回报数据
portfolio_values = pd.Series(strat.portfolio_values, index=pd.to_datetime(strat.dates))
portfolio_daily_returns = portfolio_values.pct_change().fillna(0)
portfolio_cum_returns = (1 + portfolio_daily_returns).cumprod() - 1
index_daily_returns = index_data['close'].pct_change().fillna(0)
index_cum_returns = (1 + index_daily_returns).cumprod() - 1

# 对齐日期
common_index = portfolio_daily_returns.index.intersection(index_daily_returns.index)
portfolio_daily_returns = portfolio_daily_returns.loc[common_index]
portfolio_cum_returns = portfolio_cum_returns.loc[common_index]
index_daily_returns = index_daily_returns.loc[common_index]
index_cum_returns = index_cum_returns.loc[common_index]

# 生成 Excel 表格
returns_df = pd.DataFrame({
    'Date': common_index,
    'Portfolio_Daily_Return': portfolio_daily_returns,
    'CSI500_Daily_Return': index_daily_returns
})
excel_file = "daily_returns_comparison.xlsx"
returns_df.to_excel(excel_file, index=False, sheet_name='Daily Returns')

# 格式化 Excel
wb = openpyxl.load_workbook(excel_file)
ws = wb['Daily Returns']
ws['A1'] = '日期'
ws['B1'] = '组合日收益率'
ws['C1'] = '中证500日收益率'

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.number_format = 'YYYY-MM-DD'

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
    for cell in row:
        cell.number_format = '0.00%'

wb.save(excel_file)

# 生成累计收益率图表
fig = go.Figure()
if not portfolio_cum_returns.empty and portfolio_cum_returns.isna().sum() < len(portfolio_cum_returns):
    fig.add_trace(
        go.Scatter(
            x=portfolio_cum_returns.index,
            y=portfolio_cum_returns,
            mode='lines',
            name='Portfolio Cumulative Return',
            line=dict(color='blue', width=2),
            connectgaps=True
        )
    )
    fig.add_annotation(
        x=portfolio_cum_returns.index[-1],
        y=portfolio_cum_returns.iloc[-1],
        text=f"Portfolio: {portfolio_cum_returns.iloc[-1]*100:.2f}%",
        showarrow=True,
        arrowhead=2,
        ax=20,
        ay=-30,
        font=dict(color='blue')
    )

fig.add_trace(
    go.Scatter(
        x=index_cum_returns.index,
        y=index_cum_returns,
        mode='lines',
        name='CSI 500 Cumulative Return',
        line=dict(color='red', dash='dash', width=2),
        connectgaps=True
    )
)
fig.add_annotation(
    x=index_cum_returns.index[-1],
    y=index_cum_returns.iloc[-1],
    text=f"CSI 500: {index_cum_returns.iloc[-1]*100:.2f}%",
    showarrow=True,
    arrowhead=2,
    ax=20,
    ay=30,
    font=dict(color='red')
)

fig.update_layout(
    title="指数增强策略累计收益率 (2020-01-04 至 2022-12-30)",
    height=500,
    showlegend=True,
    legend_title_text="Portfolio vs. CSI 500",
    xaxis=dict(
        title="日期",
        type='date',
        tickformat="%Y-%m-%d",
        range=[pd.to_datetime("2020-01-04"), pd.to_datetime("2022-12-30")]
    ),
    yaxis=dict(title="累计收益率 (%)", tickformat=".2%"),
    hovermode="x unified"
)

fig.write_html("cumulative_returns.html")

# 性能指标
total_return = strat.analyzers.returns.get_analysis()['rtot']
days = (portfolio_values.index[-1] - portfolio_values.index[0]).days
annualized_return = (1 + total_return) ** (252 / days) - 1
num_trades = strat.analyzers.trades.get_analysis()['total']['total']
max_drawdown = strat.analyzers.drawdown.get_analysis()['max']['drawdown'] / 100

print(f"最终资金: {cerebro.broker.getvalue():.2f}")
print(f"总回报: {total_return*100:.2f}%")
print(f"年化回报: {annualized_return*100:.2f}%")
print(f"最大回撤: {-max_drawdown*100:.2f}%")
print(f"总交易次数: {num_trades}")
print(f"日收益率比较已导出至: {excel_file}")
print(f"累计收益率图表已生成: cumulative_returns.html")