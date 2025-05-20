import akshare as ak
import pandas as pd
import backtrader as bt
import logging
import numpy as np
import talib
import os
import time
from tqdm import tqdm
from datetime import datetime, timedelta
from joblib import Parallel, delayed
import openpyxl
from openpyxl.styles import numbers
import plotly.graph_objects as go

# 设置日志，指定 utf-8 编码
logger = logging.getLogger()
logger.setLevel(logging.INFO)

log_file = 'strategy.log'
if os.path.exists(log_file):
    os.remove(log_file)
    logging.debug(f"Deleted existing log file: {log_file}")

file_handler = logging.FileHandler(log_file, encoding='utf-8')
file_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# 创建目录
os.makedirs("data/daily/cleaned", exist_ok=True)
os.makedirs("data/pe_pb/cleaned", exist_ok=True)
os.makedirs("data/financial/cleaned", exist_ok=True)

# 数据层：获取中证500成分股和数据
def get_stock_code(symbol, exchange):
    """将6位股票代码转换为带交易所前缀的格式"""
    if exchange == "上海证券交易所":
        return f"sh{symbol}"
    elif exchange == "深圳证券交易所":
        return f"sz{symbol}"
    return symbol

def getData():
    start_date = "20200104"
    end_date = "20221230"
    
    # 检查缓存
    cache_file = "data/csi500_data_cache.pkl"
    if os.path.exists(cache_file):
        logging.info("Loading cached CSI 500 data")
        cached_data = pd.read_pickle(cache_file)
        return cached_data['symbols'], cached_data['index_weights']
    
    # 获取中证500成分股
    logging.info("Fetching CSI 500 constituents")
    csi500_constituents = ak.index_stock_cons_weight_csindex(symbol="000905")
    if csi500_constituents is None or csi500_constituents.empty:
        logging.error("CSI 500 constituents data is empty or None")
        raise ValueError("CSI 500 constituents data is empty or None")
    csi500_constituents['stock_code'] = csi500_constituents.apply(
        lambda x: get_stock_code(x['成分券代码'], x['交易所']), axis=1
    )
    csi500_constituents['日期'] = pd.to_datetime(csi500_constituents['日期'])
    csi500_constituents.to_csv("data/csi500_weights.csv", encoding='utf-8')
    logging.info(f"CSI 500 constituents saved, rows: {len(csi500_constituents)}")
    time.sleep(2)
    
    symbols = csi500_constituents['stock_code'].unique().tolist()[:12]
    logging.info(f"Total unique symbols: {len(symbols)}")
    
    # 获取指数数据
    logging.info("Fetching CSI 500 index data")
    index_data = ak.index_zh_a_hist(symbol="000905", period="daily", start_date=start_date, end_date=end_date)
    if index_data is None or index_data.empty:
        logging.error("Index data is empty or None")
        raise ValueError("Index data is empty or None")
    index_data['date'] = pd.to_datetime(index_data['日期'])
    index_data = index_data.rename(columns={
        '开盘': 'open', '收盘': 'close', '最高': 'high', '最低': 'low', '成交量': 'volume'
    })
    index_data.set_index('date', inplace=True)
    index_data['open'] = index_data['open'].ffill().bfill()
    index_data['high'] = index_data['high'].ffill().bfill()
    index_data['low'] = index_data['low'].ffill().bfill()
    index_data['close'] = index_data['close'].ffill().bfill()
    index_data['volume'] = index_data['volume'].replace([np.inf, -np.inf], 0).fillna(0)
    index_data = index_data.sort_index()
    if index_data['close'].isna().any() or (index_data['close'] <= 0).any():
        logging.error(f"Index data contains invalid close prices: NaN={index_data['close'].isna().sum()}, Zero={(index_data['close'] <= 0).sum()}")
        raise ValueError("Index data contains invalid close prices")
    index_data.index.name = 'date'
    index_data.to_csv("data/csi500_index.csv", encoding='utf-8')
    logging.info(f"Index data saved, rows: {len(index_data)}")
    time.sleep(2)

    valid_symbols = []
    for symbol in tqdm(symbols, desc="Processing stocks", unit="stock"):
        cleaned_daily_file = f"data/daily/cleaned/stock_{symbol}_cleaned.csv"
        cleaned_pe_pb_file = f"data/pe_pb/cleaned/stock_{symbol}_pe_pb_cleaned.csv"
        cleaned_financial_file = f"data/financial/cleaned/stock_{symbol}_financial_cleaned.csv"

        if (os.path.exists(cleaned_daily_file) and os.path.getsize(cleaned_daily_file) > 0 and
            os.path.exists(cleaned_pe_pb_file) and os.path.getsize(cleaned_pe_pb_file) > 0 and
            os.path.exists(cleaned_financial_file) and os.path.getsize(cleaned_financial_file) > 0):
            try:
                stock_data = pd.read_csv(cleaned_daily_file, parse_dates=['date'], index_col='date')
                if not (stock_data['close'].isna().any() or (stock_data['close'] <= 0).any()):
                    valid_symbols.append(symbol)
                    continue
            except Exception as e:
                logging.warning(f"股票 {symbol} 读取现有CSV失败: {str(e)}, 将重新生成")
                # Proceed to regenerate the CSV

        symbol_code = symbol[2:]
        try:
            # 检查股票有效性
            logging.info(f"Checking validity of stock {symbol}")
            stock_info = ak.stock_individual_info_em(symbol=symbol_code)
            if stock_info is None or stock_info.empty:
                logging.warning(f"股票 {symbol} 无有效信息，可能已退市或数据不可用")
                continue

            # 获取日K线数据
            logging.info(f"Fetching daily data for {symbol}")
            time.sleep(1)
            stock_data = ak.stock_zh_a_daily(symbol=symbol, adjust="qfq", start_date=start_date, end_date=end_date)
            if stock_data is None or stock_data.empty:
                logging.warning(f"股票 {symbol} 日K线数据为空或返回 None")
                continue
            stock_data['date'] = pd.to_datetime(stock_data['date'])
            stock_data = stock_data.rename(columns={
                'open': 'open', 'high': 'high', 'low': 'low', 'close': 'close', 'volume': 'volume'
            })
            stock_data.set_index('date', inplace=True)
            stock_data['volume'] = stock_data['volume'].replace([np.inf, -np.inf], 0).fillna(0)
            
            # 增强清洗
            stock_data['close'] = stock_data['close'].ffill().bfill()
            if stock_data['close'].isna().any() or (stock_data['close'] <= 0).any():
                mean_close = stock_data['close'].mean()
                if pd.isna(mean_close) or mean_close <= 0:
                    logging.warning(f"股票 {symbol} 无法计算有效均价")
                    continue
                stock_data['close'] = stock_data['close'].fillna(mean_close)
                if stock_data['close'].isna().any() or (stock_data['close'] <= 0).any():
                    logging.warning(f"股票 {symbol} 日K线数据填充后仍无效")
                    continue

            # 获取PE/PB数据
            logging.info(f"Fetching PE/PB data for {symbol}")
            time.sleep(1)
            pe_pb_data = ak.stock_a_indicator_lg(symbol=symbol_code)
            if pe_pb_data is None or pe_pb_data.empty:
                logging.warning(f"股票 {symbol} PE/PB 数据为空或返回 None")
                continue
            pe_pb_data['date'] = pd.to_datetime(pe_pb_data['trade_date'])
            pe_pb_data.set_index('date', inplace=True)
            pe_pb_data = pe_pb_data[(pe_pb_data.index >= pd.to_datetime(start_date)) & (pe_pb_data.index <= pd.to_datetime(end_date))]
            
            # 获取财务数据
            logging.info(f"Fetching financial data for {symbol}")
            time.sleep(1)
            financial_data = ak.stock_financial_analysis_indicator(symbol=symbol_code, start_year="2020")
            if financial_data is None or financial_data.empty:
                logging.warning(f"股票 {symbol} 财务数据为空或返回 None")
                continue
            financial_data['date'] = pd.to_datetime(financial_data['日期'])
            financial_data.set_index('date', inplace=True)
            financial_data = financial_data[(financial_data.index >= pd.to_datetime(start_date)) & (financial_data.index <= pd.to_datetime(end_date))]
            
            # 清洗数据
            base_index = pd.date_range(start=start_date, end=end_date, freq='D')
            stock_data = stock_data.reindex(base_index).ffill().bfill()
            pe_pb_data = pe_pb_data.reindex(base_index).ffill().bfill()
            financial_data = financial_data.reindex(base_index).ffill().bfill()

            stock_data['volume'] = stock_data['volume'].fillna(0)
            stock_data['close'] = stock_data['close'].fillna(stock_data['close'].mean())
            stock_data['open'] = stock_data['open'].fillna(stock_data['close'])
            stock_data['high'] = stock_data['high'].fillna(stock_data['close'])
            stock_data['low'] = stock_data['low'].fillna(stock_data['close'])
            pe_pb_data['pe'] = pe_pb_data['pe'].fillna(pe_pb_data['pe'].mean()).fillna(20)
            pe_pb_data['pb'] = pe_pb_data['pb'].fillna(pe_pb_data['pb'].mean()).fillna(2)
            financial_data['净资产收益率(%)'] = financial_data['净资产收益率(%)'].fillna(financial_data['净资产收益率(%)'].mean()).fillna(0)
            financial_data['主营业务收入增长率(%)'] = financial_data['主营业务收入增长率(%)'].fillna(financial_data['主营业务收入增长率(%)'].mean()).fillna(0)
            financial_data['资产负债率(%)'] = financial_data['资产负债率(%)'].fillna(financial_data['资产负债率(%)'].mean()).fillna(50)

            if (stock_data['close'].isna().any() or (stock_data['close'] <= 0).any() or
                pe_pb_data['pe'].isna().any() or pe_pb_data['pb'].isna().any() or
                financial_data['净资产收益率(%)'].isna().any() or
                financial_data['主营业务收入增长率(%)'].isna().any() or
                financial_data['资产负债率(%)'].isna().any()):
                logging.warning(f"股票 {symbol} 清洗后数据仍包含无效值")
                continue

            # Set index name to ensure 'date' column in CSV
            stock_data.index.name = 'date'
            pe_pb_data.index.name = 'date'
            financial_data.index.name = 'date'

            stock_data.to_csv(cleaned_daily_file, encoding='utf-8')
            pe_pb_data.to_csv(cleaned_pe_pb_file, encoding='utf-8')
            financial_data.to_csv(cleaned_financial_file, encoding='utf-8')
            valid_symbols.append(symbol)
        except Exception as e:
            logging.error(f"股票 {symbol} 数据处理失败: {str(e)}")
            continue
    
    # 缓存数据
    pd.to_pickle({'symbols': valid_symbols, 'index_weights': csi500_constituents}, cache_file)
    return valid_symbols, csi500_constituents

symbols, index_weights = getData()

# 因子层：计算多因子
column_mapping = {
    "净资产收益率(%)": "roe",
    "主营业务收入增长率(%)": "revenue_growth",
    "资产负债率(%)": "debt_ratio",
}

def calculate_factors(symbol):
    try:
        stock_data = pd.read_csv(f"data/daily/cleaned/stock_{symbol}_cleaned.csv", parse_dates=['date'], index_col='date')
        pe_pb_data = pd.read_csv(f"data/pe_pb/cleaned/stock_{symbol}_pe_pb_cleaned.csv", parse_dates=['date'], index_col='date')
        financial_data = pd.read_csv(f"data/financial/cleaned/stock_{symbol}_financial_cleaned.csv", parse_dates=['date'], index_col='date')
        
        if (stock_data['close'].isna().all() or (stock_data['close'] <= 0).all() or 
            'pe' not in pe_pb_data.columns or pe_pb_data['pe'].isna().all() or 
            'pb' not in pe_pb_data.columns or pe_pb_data['pb'].isna().all() or 
            '净资产收益率(%)' not in financial_data.columns or financial_data['净资产收益率(%)'].isna().all() or 
            '主营业务收入增长率(%)' not in financial_data.columns or financial_data['主营业务收入增长率(%)'].isna().all() or
            '资产负债率(%)' not in financial_data.columns or financial_data['资产负债率(%)'].isna().all()):
            logging.warning(f"股票 {symbol} 数据不完整")
            return pd.DataFrame()
        
        factors = pd.DataFrame(index=stock_data.index)
        financial_data = financial_data.rename(columns=column_mapping)
        
        factors['pe'] = pe_pb_data['pe'].fillna(pe_pb_data['pe'].mean()).fillna(20)
        factors['pb'] = pe_pb_data['pb'].fillna(pe_pb_data['pb'].mean()).fillna(2)
        factors['roe'] = financial_data['roe'].fillna(financial_data['roe'].mean()).fillna(0)
        factors['revenue_growth'] = financial_data['revenue_growth'].fillna(financial_data['revenue_growth'].mean()).fillna(0)
        factors['debt_ratio'] = financial_data['debt_ratio'].fillna(financial_data['debt_ratio'].mean()).fillna(50)
        
        factors['short_momentum'] = stock_data['close'].pct_change(20).fillna(0)
        factors['reversal'] = -stock_data['close'].pct_change(5).fillna(0)
        factors['volatility'] = stock_data['close'].pct_change().rolling(60).std().fillna(0)
        factors['rsi'] = pd.Series(talib.RSI(stock_data['close'], timeperiod=14), index=stock_data.index).fillna(50)
        factors['low_vol'] = -factors['volatility']
        
        factor_columns = ['pe', 'pb', 'roe', 'revenue_growth', 'debt_ratio', 
                         'short_momentum', 'reversal', 'volatility', 'rsi', 'low_vol']
        for col in factor_columns:
            mean = factors[col].mean()
            std = factors[col].std()
            factors[col] = factors[col] - mean if pd.isna(std) or std == 0 else (factors[col] - mean) / std
        
        factors['score'] = (
            -factors['pe'] - factors['pb'] +
            factors['roe'] + factors['revenue_growth'] -
            factors['debt_ratio'] +
            0.5 * factors['short_momentum'] + 0.5 * factors['reversal'] -
            factors['volatility'] - factors['rsi'] +
            factors['low_vol']
        )
        factors['score'] = factors['score'].where(factors['volatility'] < factors['volatility'].quantile(0.75), np.nan)
        factors['symbol'] = symbol
        
        return factors
    except Exception as e:
        logging.error(f"股票 {symbol} 因子计算失败: {str(e)}")
        return pd.DataFrame()

# 并行计算因子
logging.info("Calculating factors in parallel")
# all_factors_list = Parallel(n_jobs=-1, verbose=10)(delayed(calculate_factors)(symbol) for symbol in symbols)
all_factors_list = [calculate_factors(symbol) for symbol in symbols]
all_factors_list = [f for f in all_factors_list if not f.empty]
if not all_factors_list:
    raise ValueError("所有股票因子数据为空")
all_factors = pd.concat(all_factors_list, axis=0)
all_factors.to_csv("data/all_factors.csv", encoding='utf-8')
logging.info(f"All factors saved, rows: {len(all_factors)}")

# 选股函数
def select_stocks(factors, rebalance_date, index_weights, alpha=0.5):
    try:
        rebalance_date = pd.Timestamp(rebalance_date)
        available_dates = factors.index.unique()
        if rebalance_date not in available_dates:
            nearest_date = available_dates[available_dates <= rebalance_date][-1] if len(available_dates[available_dates <= rebalance_date]) > 0 else available_dates[0]
            factors = factors[factors.index == nearest_date]
        else:
            factors = factors[factors.index == rebalance_date]
        
        if factors.empty:
            logging.warning(f"因子数据为空，日期: {rebalance_date}")
            return {}
        
        factors = factors.sort_values('score', ascending=False)
        num_stocks = min(100, len(factors))
        selected = factors.head(num_stocks)['symbol'].tolist()
        
        index_weights = index_weights[index_weights['日期'] <= rebalance_date].tail(1)
        index_weights_dict = {row['stock_code']: row['权重']/100 for _, row in index_weights.iterrows()}
        
        factor_weights = {stock: 1.0/num_stocks for stock in selected}
        volatilities = {}
        for stock in selected:
            symbol_data = pd.read_csv(f"data/daily/cleaned/stock_{stock}_cleaned.csv", parse_dates=['date'], index_col='date')
            window = symbol_data['close'][:rebalance_date].tail(20)
            volatilities[stock] = window.pct_change().std() if len(window) >= 10 else 0.1
        
        final_weights = {}
        total_weight = 0
        max_weight = 0.10
        vol_sum = sum(1 / v for v in volatilities.values())
        for stock in selected:
            index_weight = index_weights_dict.get(stock, 0)
            factor_weight = factor_weights.get(stock, 0)
            vol_weight = (1 / volatilities[stock]) / vol_sum
            weight = (1 - alpha) * index_weight + alpha * vol_weight
            final_weights[stock] = min(weight, max_weight)
            total_weight += final_weights[stock]
        
        if total_weight > 0:
            final_weights = {k: v/total_weight for k, v in final_weights.items()}
        return final_weights
    except Exception as e:
        logging.error(f"选股失败，日期: {rebalance_date}, 错误: {str(e)}")
        return {}

# 指数增强策略
class IndexEnhanceStrategy(bt.Strategy):
    params = (
        ('alpha', 0.5),
        ('stop_loss_stock', 0.10),
        ('stop_loss_portfolio', 0.20),
    )

    def __init__(self):
        logging.info("Initializing IndexEnhanceStrategy")
        self.all_factors = pd.read_csv("data/all_factors.csv", parse_dates=['date'], index_col='date')
        self.index_weights = pd.read_csv("data/csi500_weights.csv", parse_dates=['日期'])
        self.weights = {}
        self.prev_weights = {}
        self.portfolio_values = []
        self.dates = []
        self.initial_value = self.broker.getvalue()
        logging.info(f"Initial portfolio value: {self.initial_value}")

    def next(self):
        dt = pd.Timestamp(self.datas[0].datetime.date(0))
        portfolio_value = self.broker.getvalue()
        self.dates.append(dt)
        self.portfolio_values.append(portfolio_value)

        if pd.isna(self.datas[0].close[0]) or self.datas[0].close[0] <= 0:
            logging.warning(f"Invalid index close price: {self.datas[0].close[0]} at {dt}")
            return

        if (self.initial_value - portfolio_value) / self.initial_value > self.params.stop_loss_portfolio:
            for data in self.datas[1:]:
                self.close(data=data)
            logging.info(f"Portfolio stop-loss triggered at {dt}")
            return

        for data in self.datas[1:]:
            position = self.getposition(data)
            if position.size > 0:
                entry_price = position.price
                current_price = data.close[0]
                if pd.isna(current_price) or current_price <= 0:
                    logging.warning(f"Invalid stock close price for {data._name}: {current_price} at {dt}")
                    continue
                if (entry_price - current_price) / entry_price > self.params.stop_loss_stock:
                    self.close(data=data)

        new_weights = select_stocks(self.all_factors, dt, self.index_weights, self.params.alpha)
        if not new_weights:
            return

        self.weights = new_weights
        self.prev_weights = new_weights.copy()

        for data in self.datas[1:]:
            symbol = data._name
            target_weight = self.weights.get(symbol, 0)
            current_pos = self.getposition(data).size
            price = data.close[0]
            if target_weight > 0:
                if pd.isna(price) or price <= 0:
                    logging.warning(f"Skipping trade for {symbol} due to invalid price: {price} at {dt}")
                    continue
                try:
                    target_shares = int(target_weight * portfolio_value / price / 100) * 100
                    if target_shares > current_pos:
                        self.buy(data=data, size=target_shares - current_pos)
                    elif target_shares < current_pos:
                        self.sell(data=data, size=current_pos - target_shares)
                except (ValueError, TypeError) as e:
                    logging.error(f"Error calculating target_shares for {symbol} at {dt}: {str(e)}")
                    continue

# 回测
cerebro = bt.Cerebro(stdstats=False)
index_data = pd.read_csv("data/csi500_index.csv", parse_dates=['date'], index_col='date')
if index_data.empty or 'close' not in index_data.columns or index_data['close'].isna().any() or (index_data['close'] <= 0).any():
    logging.error("Invalid index data")
    raise ValueError("Invalid index data")
index_data = index_data.sort_index()
cerebro.adddata(bt.feeds.PandasData(
    dataname=index_data,
    name="csi500",
    datetime=None,
    open='open',
    high='high',
    low='low',
    close='close',
    volume='volume',
    fromdate=pd.to_datetime("2020-01-04"),
    todate=pd.to_datetime("2022-12-30")
))

data_feed_count = 1
for symbol in symbols:
    try:
        data = pd.read_csv(f"data/daily/cleaned/stock_{symbol}_cleaned.csv", parse_dates=['date'], index_col='date')
        if data.empty or 'close' not in data.columns or data['close'].isna().any() or (data['close'] <= 0).any():
            logging.warning(f"Skipping {symbol} due to invalid data")
            continue
        data = data.sort_index()
        data = data.reindex(index_data.index).ffill().bfill()
        data['open'] = data['open'].fillna(data['close'])
        data['high'] = data['high'].fillna(data['close'])
        data['low'] = data['low'].fillna(data['close'])
        data['volume'] = data['volume'].fillna(0)
        if data['close'].isna().any() or (data['close'] <= 0).any():
            logging.warning(f"Data for {symbol} still invalid after filling")
            continue
        cerebro.adddata(bt.feeds.PandasData(
            dataname=data,
            name=symbol,
            datetime=None,
            open='open',
            high='high',
            low='low',
            close='close',
            volume='volume',
            fromdate=pd.to_datetime("2020-01-04"),
            todate=pd.to_datetime("2022-12-30")
        ))
        data_feed_count += 1
    except Exception as e:
        logging.error(f"加载股票 {symbol} 数据失败: {str(e)}")
        continue

logging.info(f"Total data feeds loaded: {data_feed_count}")
if data_feed_count <= 1:
    raise ValueError("No valid stock data feeds loaded")

cerebro.addstrategy(IndexEnhanceStrategy)
cerebro.broker.setcash(1000000.0)
cerebro.broker.setcommission(commission=0.001)
cerebro.addsizer(bt.sizers.FixedSize, stake=100)
cerebro.addanalyzer(bt.analyzers.Returns, _name='returns')
cerebro.addanalyzer(bt.analyzers.TradeAnalyzer, _name='trades')
cerebro.addanalyzer(bt.analyzers.DrawDown, _name='drawdown')

logging.info("Starting backtest")
results = cerebro.run(runonce=True, preload=True, exactbars=0)
strat = results[0]
logging.info("Backtest completed")

# 处理回报数据
portfolio_values = pd.Series(strat.portfolio_values, index=pd.to_datetime(strat.dates))
portfolio_daily_returns = portfolio_values.pct_change().fillna(0)
portfolio_cum_returns = (1 + portfolio_daily_returns).cumprod() - 1
index_daily_returns = index_data['close'].pct_change().fillna(0)
index_cum_returns = (1 + index_daily_returns).cumprod() - 1

# 对齐日期
common_index = portfolio_daily_returns.index.intersection(index_daily_returns.index)
portfolio_daily_returns = portfolio_daily_returns.loc[common_index]
portfolio_cum_returns = portfolio_cum_returns.loc[common_index]
index_daily_returns = index_daily_returns.loc[common_index]
index_cum_returns = index_cum_returns.loc[common_index]

# 生成 Excel 表格
returns_df = pd.DataFrame({
    'Date': common_index,
    'Portfolio_Daily_Return': portfolio_daily_returns,
    'CSI500_Daily_Return': index_daily_returns
})
excel_file = "daily_returns_comparison.xlsx"
returns_df.to_excel(excel_file, index=False, sheet_name='Daily Returns')

# 格式化 Excel
wb = openpyxl.load_workbook(excel_file)
ws = wb['Daily Returns']
ws['A1'] = '日期'
ws['B1'] = '组合日收益率'
ws['C1'] = '中证500日收益率'

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.number_format = 'YYYY-MM-DD'

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
    for cell in row:
        cell.number_format = '0.00%'

wb.save(excel_file)
logging.info(f"Daily returns exported to {excel_file}")

# 生成累计收益率图表
fig = go.Figure()
if not portfolio_cum_returns.empty and portfolio_cum_returns.isna().sum() < len(portfolio_cum_returns):
    fig.add_trace(
        go.Scatter(
            x=portfolio_cum_returns.index,
            y=portfolio_cum_returns,
            mode='lines',
            name='Portfolio Cumulative Return',
            line=dict(color='blue', width=2),
            connectgaps=True
        )
    )
    fig.add_annotation(
        x=portfolio_cum_returns.index[-1],
        y=portfolio_cum_returns.iloc[-1],
        text=f"Portfolio: {portfolio_cum_returns.iloc[-1]*100:.2f}%",
        showarrow=True,
        arrowhead=2,
        ax=20,
        ay=-30,
        font=dict(color='blue')
    )
else:
    logging.warning("portfolio_cum_returns 为空或全为 NaN，跳过累计收益率组合曲线")

fig.add_trace(
    go.Scatter(
        x=index_cum_returns.index,
        y=index_cum_returns,
        mode='lines',
        name='CSI 500 Cumulative Return',
        line=dict(color='red', dash='dash', width=2),
        connectgaps=True
    )
)
fig.add_annotation(
    x=index_cum_returns.index[-1],
    y=index_cum_returns.iloc[-1],
    text=f"CSI 500: {index_cum_returns.iloc[-1]*100:.2f}%",
    showarrow=True,
    arrowhead=2,
    ax=20,
    ay=30,
    font=dict(color='red')
)

fig.update_layout(
    title="指数增强策略累计收益率 (2020-01-04 至 2022-12-30)",
    height=500,
    showlegend=True,
    legend_title_text="Portfolio vs. CSI 500",
    xaxis=dict(
        title="日期",
        type='date',
        tickformat="%Y-%m-%d",
        range=[pd.to_datetime("2020-01-04"), pd.to_datetime("2022-12-30")]
    ),
    yaxis=dict(title="累计收益率 (%)", tickformat=".2%"),
    hovermode="x unified"
)

fig.write_html("cumulative_returns.html")
logging.info("Cumulative returns chart saved to cumulative_returns.html")

# 性能指标
total_return = strat.analyzers.returns.get_analysis()['rtot']
days = (portfolio_values.index[-1] - portfolio_values.index[0]).days
annualized_return = (1 + total_return) ** (252 / days) - 1
num_trades = strat.analyzers.trades.get_analysis()['total']['total']
max_drawdown = strat.analyzers.drawdown.get_analysis()['max']['drawdown'] / 100

print(f"最终资金: {cerebro.broker.getvalue():.2f}")
print(f"总回报: {total_return*100:.2f}%")
print(f"年化回报: {annualized_return*100:.2f}%")
print(f"最大回撤: {-max_drawdown*100:.2f}%")
print(f"总交易次数: {num_trades}")
print(f"日收益率比较已导出至: {excel_file}")
print(f"累计收益率图表已生成: cumulative_returns.html")